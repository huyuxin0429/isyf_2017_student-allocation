"""
ISYF 2017 student allocation program
Takes in list of student names and 2 preferences
Ouputs in Results.xlsx

By:
Hu Yu Xin

"""
import csv
import openpyxl
import os
import random
from openpyxl import Workbook,load_workbook
#Initialisation

#Output initialisation
wboutput = Workbook()   #Creating output workbook
wsoutput = wboutput.active #Creating output worksheet

wsoutput["A1"] = "Student"   #Creating output headers
wsoutput["B1"] = "Class Allocation"
wsoutput["C1"] = "Dialogue Allocation"
#Output initialisation end

#Classes array initialisation
classes =[]                 #Definiting array for class allocation
for x in range(0,4):
    classes.append([0])     #Number of student in each class
    
classes[0].append(48)       #Second number determines maximum class size
classes[1].append(24)
classes[2].append(24)
classes[3].append(26)
#Classes array initialisation end

dialogue = []
for x in range(0,5):
    dialogue.append([0])
dialogue[0].append(80)
dialogue[1].append(80)
dialogue[2].append(80)
dialogue[3].append(80)
dialogue[4].append(80)

#debug = bool(int(input("Debug view? Enter 1 for yes, 0 for no ")))           #Debug view toggle
debug = True

allocationIndex = ["Bio1","Bio2","Phy","Comp"] #Index for each class
"""Reference index for classes
0: Bio1
1: Bio2
2: Phy
3: Comp
"""
dialogueIndex = ["Math","Com","Phy","Bio","Chem"] #Index for each class
"""Reference index for dialogue
0: Math
1: Com
2: Phy
3: Bio
4: Chem
"""
# Input initialisation
#filename = input("Name of source file? Write in full Eg. Student_Choice.xlsx ") #User input source file name
filename = "1.xlsx"
file_path = os.path.join(os.getcwd()+"/"+filename)
inputworkbook = load_workbook(file_path)
inputworksheet = inputworkbook.active

#customise = bool(int(input(
"""Customise input data range?
Default options:
Header(Title rows, no data): 1 row
Name colomn: 2
Prefernce 1 colomm: 8
Preference 2 colomn: 9
Enter 1 for yes, 0 for no """
#)))
customise = False
if customise:
    header = int(input("Enter header rows:"))
    namecol =int(input("Enter name colomn no.:"))
    choice1col =int(input("Enter choice 1 colomn no.:"))
    choice2col =int(input("Enter choice 2 colomn no.:"))
else:
    header = 1
    namecol = 2
    choice1col = 8
    choice2col = 9
    print(header,namecol,choice1col,choice2col)
# Input initialisation end

#End Initialisation

def stucount(worksheet): #Counting and returning total no. of students by finding no. of lines to first null value
    x = header+1 # row start at header row
    count = 0
    while worksheet.cell(row = x, column = namecol).value != None: 
        count += 1
        x += 1
    print("Total no. of students:",count)
    return count
"""
Preference Choices:
0: math
1: phy
2: bio
3: chem
"""
def prefconversion(preference1,array):
    if isinstance(array,int)or isinstance(array,str):
        conv = [array]
        return prefconversion(preference1,conv)
    if compref(preference1) in array:
        return compref(preference1)
    else:
        return randomfrmarray(array,preference1)
def compref(preference):
    compref={"Physics":"Mathematics","Chemistry":"Biology","Mathematics":"Physics","Biology":"Chemistry"}
    return compref[preference]
def prefnumber(var):
    table1 = {"Physics":1,"Chemistry":3,"Mathematics":0,"Biology":2}
    if isinstance(var,str):
        return table1[var]
    if isinstance(var,int):
        table2 = {}
        for key in table1:
            table2[table1[key]]=key
        return table2[var]


def allocate(preference1,preference2,background):
    #Allocation selector, insert 2 preference, output allocation based on student background
    #Allocate by preference and class availibility
    if debug:
        print("Student choice:",prefnumber(preference1),prefnumber(preference2))
    allocated = False
    preference=[preference1,preference2]
    prefincre = 0 
    while prefincre <=1:
        if debug:
                if prefincre == 0:
                    print("Allocation by preference 1...",preference1,prefnumber(preference1))
                else:
                    print("Allocation by preference 2...",preference2,prefnumber(preference2))
        if preference[prefincre] == 2:#Bio choice have 2 class
            if alloclass(1):
                return 1
            elif alloclass(0):
                return 0
        elif preference[prefincre] == 1:
            if alloclass(2):
                return 2
        elif preference[prefincre] == 3:
            if alloclass(1):
                return 1
            elif alloclass(0):
                return 0
        elif preference[prefincre] == 0:
            if alloclass(3):
                return 3
        prefincre += 1
        if debug:
            if prefincre == 1:
                print("Prefernce 1 allocation failed")
            else:
                print("Prefernce 2 allocation failed")
    finaloptions = [0,1,2,3]
    for element in background:
        finaloptions.remove(element)
        print("test1")
    #for element in preference:
    #    background.remove(element)
    #    print("test2")
    for element in background:
        print("test3")
        if alloclass(element):
            return element
    for element in finaloptions:
        print("test4")
        if allocate(element):
            return element
    print("Can't allocate", student[b][0],"Max student no. exceeded?")
    return False
def alloclass(clas):
    if classes[clas][0]<classes[clas][1]: 
        classes[clas][0] +=1
        if debug:
            print(allocationIndex[clas],"class size:",classes[clas][0])
        return True
    else:
        return False
def randomfrmarray(full,choice = False):
    if isinstance(choice,list):
        for element in choice:
            full.remove(element)
        return full[random.random(0,len(full)-1)]
    else:
        return full[random.randint(0,len(full)-1)]
        
def randomallocate(full,choice):
        allocated = False
        while allocated == False:   #If both preferences are unavilible, randomly selects a class
            if debug:
                print("Assigning to random availible class...")
            randomclass = randomfrmarray([0,1,2,3],computedclass,True)
            if debug:
                print(randomclass)
            allocatedclass = classarray[randomclass]    #Randomly selecting among last  class
            if classes[allocatedclass][0]<classes[allocatedclass][1]:
                classes[allocatedclass][0]+=1
                if debug:
                        print(allocationIndex[allocatedclass],"class size:",classes[allocatedclass][0])
                allocated = True
                return allocatedclass
def allocateDialogue(preference1,array):
    #Allocation selector, insert 2 preference, output allocation
    #Allocate by preference and lecture availibility
    preference1 -= 1 #Convert from range 1-5 to 0-4
    if debug:
        print("Student choice:",preferenceIndex[preference1],preferenceIndex[preference2])
    allocated = False
    preference=[preference1,preference2]
    while allocated == False: #Allocation status flag
        computeddialogue=[0,0]
        for u in range(0,2):
            computeddialogue[u] = preference[u]
        prefincre = 0 
        while prefincre <=1:
            if debug:
                    if prefincre == 0:
                        print("Allocation by preference 1...",preference1,preferenceIndex[preference1])
                    else:
                        print("Allocation by preference 2...",preference2,preferenceIndex[preference2])
            if dialogue[computeddialogue[prefincre]][0]<dialogue[computeddialogue[prefincre]][1]:
                dialogue[computeddialogue[prefincre]][0] +=1
                if debug:
                        print(allocationIndex[computeddialogue[prefincre]],"class size:",dialogue[computeddialogue[prefincre]][0])
                allocated = True 
                return computeddialogue[prefincre]
            else:
                prefincre += 1
                if debug:
                    if prefincre == 1:
                        print("Prefernce 1 allocation failed")
                    else:
                        print("Prefernce 2 allocation failed")
        while allocated == False:   #If both preferences are unavilible, randomly selects a class
            if debug:
                print("Assigning to random availible class...")
            dialoguearray = [0,1,2,3,4]
            for x in range(0,2):
                if computeddialogue[x] == 0:
                    dialoguearray.remove(0)
                    dialoguearray.remove(1)#Removing preferences from availible classes
                else:
                    classarray.remove(computedclass[x])
            randomdialogue = random.randint(0,len(classarray)-1)
            if debug:
                print(randomdialogue)
            allocateddialogue = dialoguearray[randomdialogue]    #Randomly selecting among last  class
            if dialogue[allocateddialogue][0]<dialogue[allocateddialogue][1]:
                dialogue[allocateddialogue][0]+=1
                if debug:
                        print(lextureAllocationIndex[allocateddialogue],"class size:",dialogue[allocateddialogue][0])
                allocated = True
                return allocateddialogue
def main():
    studentarray = [] #Initialising local array database of student names and preferences
    studentcount = stucount(inputworksheet) #Retriving total no. of students
    
    for a in range(0,studentcount): #Initalising student array to accmodate for 3 values (name,preference1,preference2)
        studentarray.append([""]*3)
             
    for y in range(0,studentcount):#Copying values from input file to local database array
        studentarray[y][0]= inputworksheet.cell(row = y+1+header, column = namecol).value
        studentarray[y][1]= inputworksheet.cell(row = y+1+header, column = choice1col).value
        studentarray[y][2]= ((inputworksheet.cell(row = y+1+header, column = choice2col).value).split(", "))
        if debug:
            print("Name:",studentarray[y][0],"Choice 1:",studentarray[y][1],"Others:",studentarray[y][2])   
            print("\n")
    
    for b in range(0,studentcount): #Assigning classes to students + printing output
        if debug:
            print("Allocating student:",studentarray[b][0],"No.:",b)
        preference1 = prefnumber(studentarray[b][1])
        preference2 = prefnumber(prefconversion(studentarray[b][1],studentarray[b][2]))
        studentsubject = []
        for element in studentarray[b][2]:
            studentsubject.append(prefnumber(element))
        allocation = allocate(preference1,preference2,studentsubject)
        if isinstance(allocation,bool):
            print("Can't allocate", student[b][0],"Max student no. exceeded?")
        else:
            #dialogueAllocation = allocateDialogue (studentarray[b][1],studentsubject)
            wsoutput.cell(row=b+1+header,column=1).value = studentarray[b][0]
            wsoutput.cell(row=b+1+header,column=2).value = allocationIndex[allocation]
           # wsoutput.cell(row=b+1+header,column=3).value = dialogueIndex[dialogueAllocation]
            if debug:
                print(studentarray[b][0],"is allocated to class",allocationIndex[allocation],"\n")#,"and dialogue: ",dialogueIndex[dialogueAllocation],"\n")
    
    wboutput.save("Result.xlsx")
    print("Done!")
    
    
main()
